#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""B站关注列表清理与分析工具（仅本地运行，遵守 B站用户协议）。"""

import argparse
import datetime
import hashlib
import json
import os
import random
import re
import sys
import time
import urllib.parse
import unicodedata

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
    RESOURCE_DIR = getattr(sys, "_MEIPASS", BASE_DIR)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    RESOURCE_DIR = BASE_DIR
DATA_DIR = os.path.join(BASE_DIR, "data")
COOKIE_FILE = os.path.join(DATA_DIR, "cookies.json")
FOLLOW_FILE = os.path.join(DATA_DIR, "followings.json")
ANALYSIS_FILE = os.path.join(DATA_DIR, "analysis.json")
UNFOLLOW_STATE = os.path.join(DATA_DIR, "unfollow_state.json")
SELECTED_FILE = os.path.join(DATA_DIR, "selected_unfollow.json")
PLAN_FILE = os.path.join(BASE_DIR, "unfollow_plan.xlsx")
REPORT_FILE = os.path.join(BASE_DIR, "bili_follow_report.xlsx")
ACCOUNTS_DIR = os.path.join(DATA_DIR, "accounts")
CURRENT_ACCOUNT_FILE = os.path.join(DATA_DIR, "current_account.json")

UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)

# 分区名/关键词 -> 主分类
CATEGORY_RULES = [
    ("动画", ["动画", "番剧", "国创", "MAD", "MMD", "手书", "配音"]),
    ("音乐", ["音乐", "原创音乐", "翻唱", "演奏", "VOCALOID", "MV", "乐评"]),
    ("舞蹈", ["舞蹈", "宅舞", "街舞", "明星舞蹈", "中国舞"]),
    ("游戏", ["游戏", "单机", "电子竞技", "手机游戏", "网络游戏", "桌游", "音游", "Mugen", "电竞"]),
    ("知识", ["知识", "科学", "科普", "社科", "法律", "心理", "人文", "历史", "校园", "学习", "职业", "职场", "财经", "商业", "设计", "创意"]),
    ("科技", ["科技", "数码", "软件", "计算机", "编程", "AI", "智能", "极客", "科工"]),
    ("运动", ["运动", "篮球", "足球", "健身", "竞技体育", "体育"]),
    ("汽车", ["汽车", "赛车", "改装", "新能源", "摩托车"]),
    ("生活", ["生活", "日常", "搞笑", "亲子", "出行", "三农", "家居", "手工", "绘画", "挑战"]),
    ("动物圈", ["动物", "喵星人", "汪星人", "熊猫", "野生", "爬宠", "萌宠"]),
    ("美食", ["美食", "美食制作", "美食测评", "田园美食"]),
    ("鬼畜", ["鬼畜", "音MAD", "人力VOCALOID"]),
    ("时尚", ["时尚", "美妆", "穿搭", "仿妆", "潮流"]),
    ("资讯", ["资讯", "热点", "环球", "社会", "新闻"]),
    ("娱乐", ["娱乐", "综艺", "明星", "韩国", "杂谈"]),
    ("影视", ["影视", "影视剪辑", "影视杂谈", "短片", "预告"]),
    ("纪录片", ["纪录片", "探索", "自然", "军事"]),
    ("电影", ["电影", "华语电影", "欧美电影", "日本电影"]),
    ("电视剧", ["电视剧", "国产剧", "海外剧"]),
]


def load_json(path, default):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return default


def save_json(path, obj):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)
    for _ in range(8):
        try:
            os.replace(tmp, path)
            return
        except PermissionError:
            time.sleep(0.4)
    os.replace(tmp, path)


def current_uid():
    info = load_json(CURRENT_ACCOUNT_FILE, {})
    uid = info.get("uid")
    if not uid:
        legacy = load_json(COOKIE_FILE, {})
        uid = legacy.get("DedeUserID") or legacy.get("DedeUserID__ckMd5")
    return str(uid) if uid else ""


def account_dir(uid):
    uid = str(uid)
    d = os.path.join(ACCOUNTS_DIR, uid)
    os.makedirs(d, exist_ok=True)
    return d


def account_file(uid, name):
    return os.path.join(account_dir(uid), name)


def migrate_account(uid):
    uid = str(uid)
    d = account_dir(uid)
    for name in ("cookies.json", "followings.json", "analysis.json", "unfollow_state.json", "selected_unfollow.json"):
        dst = os.path.join(d, name)
        src = os.path.join(DATA_DIR, name)
        if os.path.exists(src) and not os.path.exists(dst):
            try:
                os.replace(src, dst)
            except OSError:
                pass
    for name in ("bili_follow_report.xlsx", "unfollow_plan.xlsx"):
        dst = os.path.join(d, name)
        src = os.path.join(BASE_DIR, name)
        if os.path.exists(src) and not os.path.exists(dst):
            try:
                os.replace(src, dst)
            except OSError:
                pass
    save_json(CURRENT_ACCOUNT_FILE, {"uid": uid})


PARTITION_FILE = os.path.join(RESOURCE_DIR, "partition_map.json")
TID_TO_CATEGORY = {int(k): v for k, v in load_json(PARTITION_FILE, {}).items()}


def now_ts():
    return int(time.time())


def to_int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def fmt_ts(ts):
    if not ts:
        return ""
    return datetime.datetime.fromtimestamp(int(ts)).strftime("%Y-%m-%d %H:%M")


def days_since(ts):
    if not ts:
        return None
    return max(0, (now_ts() - int(ts)) // 86400)


def human_days(days):
    if days is None:
        return "无"
    days = int(days)
    if days < 30:
        return f"{days}天"
    years, rem = divmod(days, 365)
    months, rem = divmod(rem, 30)
    parts = []
    if years:
        parts.append(f"{years}年")
    if months:
        parts.append(f"{months}个月")
    if rem or not parts:
        parts.append(f"{rem}天")
    return "".join(parts)


def classify_name(name, title=""):
    text = f"{name} {title}".lower()
    for cat, kws in CATEGORY_RULES:
        for kw in kws:
            if kw.lower() in text:
                return cat
    return None


def classify_videos(videos):
    counts = {}
    for v in videos:
        cat = None
        tid = v.get("typeid")
        if tid:
            try:
                cat = TID_TO_CATEGORY.get(int(tid))
            except (TypeError, ValueError):
                cat = None
        if cat is None:
            cat = classify_name(v.get("tname") or "", v.get("title") or "")
        if cat is None:
            cat = "其他/未知"
        counts[cat] = counts.get(cat, 0) + 1
    if not counts:
        return "其他/未知", "无视频"
    top = max(counts, key=lambda k: (counts[k], k))
    detail = "、".join(f"{k}×{v}" for k, v in sorted(counts.items(), key=lambda kv: -kv[1]))
    return top, detail


def video_label(v):
    tid = v.get("typeid")
    try:
        part = TID_TO_CATEGORY.get(int(tid), "未知")
    except (TypeError, ValueError):
        part = "未知"
    title = v.get("title") or "（无标题）"
    return f"【{part}】{title}"


class BiliClient:
    def __init__(self):
        self.sess = requests.Session()
        self.sess.headers.update({"User-Agent": UA, "Referer": "https://www.bilibili.com/"})
        self.uid = current_uid()
        if self.uid:
            migrate_account(self.uid)
            self.cookies = load_json(account_file(self.uid, "cookies.json"), {})
        else:
            self.cookies = load_json(COOKIE_FILE, {})
        self._apply_cookies(self.cookies)
        self.wbi_cache = None
        self.wbi_cache_time = 0

    def _apply_cookies(self, cookies):
        for k, v in cookies.items():
            if v:
                self.sess.cookies.set(k, v, domain=".bilibili.com", path="/")

    def save_cookies(self):
        uid = str(self.cookies.get("DedeUserID") or self.cookies.get("DedeUserID__ckMd5") or "")
        if uid:
            self.uid = uid
            save_json(account_file(uid, "cookies.json"), self.cookies)
            save_json(CURRENT_ACCOUNT_FILE, {"uid": uid})
        else:
            save_json(COOKIE_FILE, self.cookies)

    def account_file(self, name):
        uid = self.uid or current_uid()
        if uid:
            return account_file(uid, name)
        return os.path.join(DATA_DIR, name)

    def warm_up(self):
        # 获取匿名 buvid3，降低部分接口触发 412 的概率
        try:
            self.sess.get("https://www.bilibili.com/", timeout=10)
        except Exception:
            pass

    def get(self, url, params=None, referer=None, retries=4):
        headers = self.sess.headers.copy()
        if referer:
            headers["Referer"] = referer
        for attempt in range(retries):
            try:
                r = self.sess.get(url, params=params, headers=headers, timeout=20)
            except requests.RequestException as e:
                time.sleep(2 * (attempt + 1))
                continue
            if r.status_code == 412:
                print("  遇到 412 风控，稍后冷却重试……", file=sys.stderr)
                return {"code": -412, "message": "风控"}
            try:
                data = r.json()
            except Exception:
                data = {"code": -999, "message": "响应不是 JSON", "_text": r.text[:200]}
            if data.get("code") == -352:
                print("  遇到 -352 风控，稍后冷却重试……", file=sys.stderr)
                return {"code": -352, "message": "风控"}
            return data
        return {"code": -412, "message": "风控/重试失败"}

    def post(self, url, data=None, referer=None, retries=3):
        headers = self.sess.headers.copy()
        headers["Origin"] = "https://space.bilibili.com"
        headers["Content-Type"] = "application/x-www-form-urlencoded"
        if referer:
            headers["Referer"] = referer
        for attempt in range(retries):
            try:
                r = self.sess.post(url, data=data, headers=headers, timeout=20)
            except requests.RequestException:
                time.sleep(2 * (attempt + 1))
                continue
            if r.status_code == 412:
                print("  遇到 412 风控，稍后冷却重试……", file=sys.stderr)
                return {"code": -412, "message": "风控"}
            try:
                resp = r.json()
            except Exception:
                resp = {"code": -999, "message": "响应不是 JSON"}
            if resp.get("code") == -352:
                print("  遇到 -352 风控，稍后冷却重试……", file=sys.stderr)
                return {"code": -352, "message": "风控"}
            return resp
        return {"code": -412, "message": "风控/重试失败"}

    def login_qr(self):
        os.makedirs(DATA_DIR, exist_ok=True)
        d = self.get("https://passport.bilibili.com/x/passport-login/web/qrcode/generate",
                     referer="https://passport.bilibili.com/")
        if d.get("code") != 0 or not d.get("data"):
            print("生成登录二维码失败：", d.get("message"), file=sys.stderr)
            return False
        qr_url = d["data"]["url"]
        qrcode_key = d["data"]["qrcode_key"]
        try:
            import qrcode

            qr = qrcode.QRCode(border=2, box_size=10)
            qr.add_data(qr_url)
            qr.make(fit=True)
            img = qr.make_image()
            qr_path = os.path.join(DATA_DIR, "login_qr.png")
            os.makedirs(DATA_DIR, exist_ok=True)
            img.save(qr_path)
            print(f"二维码图片已保存：{qr_path}")
        except Exception:
            pass
        print("请用手机 B站 App 扫码登录。如果二维码图片显示不清，也可以在浏览器打开下面的链接再扫：")
        print(qr_url)
        print("等待扫码……")
        while True:
            time.sleep(2)
            p = self.get("https://passport.bilibili.com/x/passport-login/web/qrcode/poll",
                         {"qrcode_key": qrcode_key},
                         referer="https://passport.bilibili.com/")
            if not p.get("data"):
                print("登录轮询失败：", p.get("message"), file=sys.stderr)
                continue
            code = p["data"].get("code")
            if code == 0:
                return self._save_login_from_cross_domain(p["data"].get("url"))
            if code == 86038:
                print("二维码已失效，请重新运行 login。", file=sys.stderr)
                return False
            if code == 86090:
                print("已扫码，请在手机上确认登录……")
            elif code == 86101:
                print("等待扫码……")

    def _save_login_from_cross_domain(self, url):
        cookies = {}
        # 优先从会话 Cookie 读取：requests 会自动跟随登录成功后的重定向并保存 Set-Cookie
        for key in ("SESSDATA", "bili_jct", "DedeUserID", "DedeUserID__ckMd5"):
            val = self.sess.cookies.get(key, domain=".bilibili.com") or self.sess.cookies.get(key)
            if val:
                cookies[key] = urllib.parse.unquote(str(val))
        # 兜底：从成功回调 URL 的查询参数里解析
        if not cookies.get("SESSDATA") and url:
            qs = urllib.parse.parse_qs(urllib.parse.urlsplit(url).query)
            for key in ("SESSDATA", "bili_jct", "DedeUserID", "DedeUserID__ckMd5"):
                val = qs.get(key)
                if val:
                    cookies[key] = urllib.parse.unquote(val[0])
        if not cookies.get("SESSDATA") and not cookies.get("bili_jct"):
            print("未找到登录凭证，已登录 Cookie 列表：",
                  [(c.name, c.domain) for c in self.sess.cookies], file=sys.stderr)
            return False
        if not cookies.get("SESSDATA") or not cookies.get("bili_jct"):
            print("登录凭证不完整，请重试。", file=sys.stderr)
            return False
        self.cookies = cookies
        self._apply_cookies(cookies)
        self.save_cookies()
        mid = self.nav_mid()
        print(f"登录成功，当前 UID：{mid}")
        return True

    def nav_mid(self):
        d = self.get("https://api.bilibili.com/x/web-interface/nav")
        if d.get("code") == 0:
            return d["data"].get("mid")
        return self.cookies.get("DedeUserID") or ""

    def _wbi_keys(self):
        if self.wbi_cache and now_ts() - self.wbi_cache_time < 3600:
            return self.wbi_cache
        d = self.get("https://api.bilibili.com/x/web-interface/nav")
        wbi = (d.get("data") or {}).get("wbi_img") or {}
        img = wbi.get("img_url", "").rsplit("/", 1)[-1].split(".")[0]
        sub = wbi.get("sub_url", "").rsplit("/", 1)[-1].split(".")[0]
        if img and sub:
            self.wbi_cache = (img, sub)
            self.wbi_cache_time = now_ts()
        return self.wbi_cache or (img, sub)

    def wbi_sign(self, params):
        mixin_tab = [46, 47, 18, 2, 53, 8, 23, 32, 15, 50, 10, 31, 58, 3, 45, 35,
                     27, 43, 5, 49, 33, 9, 42, 19, 29, 28, 14, 39, 12, 38, 41, 13,
                     37, 48, 7, 16, 24, 55, 40, 61, 26, 17, 0, 1, 60, 51, 30, 4,
                     22, 25, 54, 21, 56, 59, 6, 63, 57, 62, 11, 36, 20, 34, 44, 52]
        img, sub = self._wbi_keys()
        mixin_key = "".join((img + sub)[i] for i in mixin_tab)[:32]
        clean = {k: "".join(ch for ch in str(v) if ch not in "!'()*") for k, v in params.items()}
        clean["wts"] = now_ts()
        clean = dict(sorted(clean.items()))
        query = urllib.parse.urlencode(clean)
        clean["w_rid"] = hashlib.md5((query + mixin_key).encode()).hexdigest()
        return clean

    def fetch_followings(self, uid, delay=0.8, start_page=1, max_pages=None):
        items = []
        page = start_page
        while True:
            params = {"vmid": uid, "pn": page, "ps": 50, "order": "desc",
                      "order_type": "attention", "jsonp": "jsonp"}
            d = self.get("https://api.bilibili.com/x/relation/followings", params,
                         referer=f"https://space.bilibili.com/{uid}/fans/follow")
            if d.get("code") != 0:
                raise RuntimeError(f"拉取关注列表失败（第 {page} 页）：{d.get('message')}")
            data = d.get("data") or {}
            lst = data.get("list") or []
            items.extend(lst)
            total = data.get("total", 0)
            print(f"已拉取 {len(items)}/{total}")
            if not lst or len(items) >= total:
                break
            if max_pages and page - start_page + 1 >= max_pages:
                break
            page += 1
            time.sleep(delay + random.random() * 0.7)
        return items

    def get_card(self, mid):
        d = self.get("https://api.bilibili.com/x/web-interface/card", {"mid": mid},
                     referer=f"https://space.bilibili.com/{mid}/video")
        if d.get("code") == -404:
            return None, True
        card = (d.get("data") or {}).get("card") or {}
        return card, False

    def get_videos(self, mid, order, ps):
        params = self.wbi_sign({"mid": mid, "ps": ps, "pn": 1, "order": order})
        d = self.get("https://api.bilibili.com/x/space/wbi/arc/search", params,
                     referer=f"https://space.bilibili.com/{mid}/video")
        if d.get("code") != 0:
            return None, d.get("code"), d.get("message"), 0
        vlist = (((d.get("data") or {}).get("list") or {}).get("vlist")) or []
        total = (((d.get("data") or {}).get("page") or {}).get("count")) or len(vlist)
        out = []
        for v in vlist:
            out.append({
                "title": v.get("title", ""),
                "play": v.get("play", 0),
                "tname": v.get("tname") or "",
                "typeid": v.get("typeid", 0),
                "created": v.get("created"),
                "bvid": v.get("bvid", ""),
            })
        return out, 0, "", total

    def get_latest_dynamic(self, mid):
        d = self.get("https://api.bilibili.com/x/polymer/web-dynamic/v1/feed/space",
                     {"host_mid": mid, "timezone_offset": -480, "features": "itemOpusStyle"},
                     referer=f"https://space.bilibili.com/{mid}/dynamic")
        if d.get("code") != 0:
            return None, d.get("code"), d.get("message")
        items = (d.get("data") or {}).get("items") or []
        for item in items:
            modules = item.get("modules") or {}
            pub_ts = (modules.get("module_author") or {}).get("pub_ts")
            if pub_ts:
                major = (modules.get("module_dynamic") or {}).get("major") or {}
                typ = major.get("type", "DYNAMIC")
                text = ""
                if typ == "MAJOR_TYPE_ARCHIVE":
                    text = (major.get("archive") or {}).get("title") or ""
                elif typ == "MAJOR_TYPE_OPUS":
                    opus = major.get("opus") or {}
                    title = opus.get("title") or ""
                    summary = opus.get("summary") or {}
                    text = title or (summary.get("text") if isinstance(summary, dict) else "")
                if not text:
                    desc = (modules.get("module_dynamic") or {}).get("desc") or {}
                    text = (desc.get("text") if isinstance(desc, dict) else desc) or ""
                return pub_ts, typ, str(text or "").strip()[:80]
        return None, 0, ""

    def unfollow(self, mid):
        csrf = self.cookies.get("bili_jct", "")
        data = {"fid": mid, "act": 2, "re_src": 11, "jsonp": "jsonp", "csrf": csrf}
        return self.post("https://api.bilibili.com/x/relation/modify", data,
                         referer=f"https://space.bilibili.com/{mid}/video")


def ensure_login(client):
    client.warm_up()
    mid = client.nav_mid()
    if not mid:
        print("尚未登录，请先运行：python bili_follow_manager.py login", file=sys.stderr)
        sys.exit(1)
    return str(mid)


def cmd_login(client, args):
    client.login_qr()


def cmd_logout(client, args):
    for path in (CURRENT_ACCOUNT_FILE, COOKIE_FILE):
        try:
            os.remove(path)
        except OSError:
            pass
    client.cookies = {}
    client.uid = ""
    client.sess.cookies.clear()
    print("已退出登录。")


def cmd_fetch(client, args):
    uid = ensure_login(client)
    items = client.fetch_followings(uid, delay=args.delay, start_page=args.start, max_pages=args.max_pages)
    path = client.account_file("followings.json")
    save_json(path, items)
    print(f"已保存关注列表：{len(items)} 人 → {path}")


def analyze_one(client, item, no_dynamic=False, top_n=10):
    def jitter():
        time.sleep(0.6 + random.random() * 0.7)

    mid = str(item.get("mid"))
    rec = {
        "mid": mid,
        "uname": item.get("uname", ""),
        "sign": item.get("sign", ""),
        "face": item.get("face", ""),
        "mtime": item.get("mtime", 0),
        "special": item.get("special", 0),
        "attribute": item.get("attribute", 0),
        "status": "ok",
        "done": True,
    }
    card, cancelled = client.get_card(mid)
    jitter()
    if cancelled or "已注销" in str(rec.get("uname")) or "已注销" in str(card.get("name") if card else ""):
        rec["status"] = "cancelled"
        rec["uname"] = rec["uname"] or "账号已注销"
        return rec
    if not card:
        rec["status"] = "error"
        rec["error"] = "用户卡片获取失败"
        return rec
    rec["uname"] = card.get("name") or rec["uname"]
    rec["fans"] = card.get("fans", 0)
    rec["level"] = card.get("level", 0)
    official = card.get("official") or {}
    rec["official"] = official.get("title") or ""

    top, code, msg, video_total = client.get_videos(mid, "click", top_n)
    jitter()
    if code not in (0, -404):
        rec["status"] = "error"
        rec["error"] = f"视频列表获取失败 code={code} {msg}"
        return rec
    rec["top_videos"] = top or []
    rec["video_total"] = video_total
    latest, code, msg, _ = client.get_videos(mid, "pubdate", 1)
    jitter()
    if code in (0, -404) and latest:
        rec["latest_video_ts"] = to_int(latest[0].get("created"))
        rec["latest_video_title"] = latest[0].get("title") or ""
    else:
        rec["latest_video_ts"] = None
        rec["latest_video_title"] = ""

    rec["latest_dynamic_ts"] = None
    rec["latest_dynamic_type"] = None
    rec["latest_dynamic_text"] = ""
    if not no_dynamic:
        dyn_ts, dyn_type, dyn_text = client.get_latest_dynamic(mid)
        jitter()
        if dyn_ts:
            rec["latest_dynamic_ts"] = to_int(dyn_ts)
            rec["latest_dynamic_type"] = dyn_type
            rec["latest_dynamic_text"] = dyn_text

    candidates = [ts for ts in (rec.get("latest_dynamic_ts"), rec.get("latest_video_ts")) if ts]
    rec["latest_post_ts"] = max(candidates) if candidates else None
    rec["category"], rec["category_detail"] = classify_videos(rec.get("top_videos") or [])
    rec["fans"] = int(rec.get("fans") or 0)
    return rec


def cmd_analyze(client, args):
    follows = load_json(client.account_file("followings.json"), [])
    if not follows:
        print("关注列表为空，请先运行 fetch。", file=sys.stderr)
        sys.exit(1)
    analysis = load_json(client.account_file("analysis.json"), {})
    start = args.start
    limit = args.limit if args.limit else (len(follows) - start)
    segment = follows[start:start + limit]
    print(f"待分析 {len(segment)} 人（从第 {start} 个开始）")
    risk_streak = 0
    for i, item in enumerate(segment):
        mid = str(item.get("mid"))
        if args.only_new and mid in analysis:
            prev = analysis[mid]
            if prev.get("done"):
                if prev.get("status") == "cancelled":
                    continue
                if prev.get("status") == "ok":
                    have = len(prev.get("top_videos") or [])
                    total = prev.get("video_total")
                    if total is not None:
                        if have >= min(args.top_n, int(total)):
                            continue
                    elif have >= args.top_n:
                        continue
        print(f"[{start + i + 1}/{len(follows)}] 分析 {item.get('uname') or mid}")
        rec = None
        is_risk = False
        for attempt in range(args.max_risk_retries + 1):
            rec = analyze_one(client, item, no_dynamic=args.no_dynamic, top_n=args.top_n)
            is_risk = rec.get("status") == "error" and any(
                s in str(rec.get("error")) for s in ("412", "-352", "风控"))
            if not is_risk or attempt >= args.max_risk_retries:
                break
            print(f"  风控，冷却 {args.cooldown} 秒后重试（第 {attempt + 1} 次）……")
            time.sleep(args.cooldown)
        rec["done"] = rec.get("status") != "error"
        rec["updated"] = now_ts()
        analysis[mid] = rec
        save_json(client.account_file("analysis.json"), analysis)
        status = rec.get("status")
        if status == "error":
            print("  错误：", rec.get("error"))
            if is_risk:
                risk_streak += 1
                if risk_streak >= 3:
                    print("连续遇到风控，已暂停分析。请等待更长时间后重新运行 analyze --only-new。")
                    break
        elif status == "cancelled":
            print("  已注销")
            risk_streak = 0
        else:
            print(f"  分类 {rec.get('category')}，粉丝 {rec.get('fans')}")
            if rec.get("latest_video_title"):
                print(f"    最新视频：{rec.get('latest_video_title')}")
            if rec.get("latest_dynamic_ts"):
                dyn_label = rec.get("latest_dynamic_text") or rec.get("latest_dynamic_type") or "动态"
                print(f"    最新动态：{dyn_label}")
            risk_streak = 0
        lo = args.delay_min if args.delay_min is not None else 1.0
        hi = args.delay_max if args.delay_max is not None else 2.0
        time.sleep(lo + random.random() * max(0.0, hi - lo))
    print(f"分析完成，结果保存在 {client.account_file('analysis.json')}")


def make_reason(rec, days_threshold):
    if rec.get("status") == "cancelled":
        return "账号已注销", "建议取关"
    if rec.get("special") == 1:
        return "特别关注（白名单）", "保留"
    if rec.get("attribute") == 6:
        return "互相关注（白名单）", "保留"
    days = days_since(rec.get("latest_post_ts"))
    if rec.get("latest_post_ts") is None:
        return "无公开投稿/动态", "建议取关"
    if days is not None and days >= days_threshold:
        reason = f"超过{days_threshold}天未更新（{human_days(days)}）"
        if int(rec.get("fans") or 0) < 1000:
            reason += "，且粉丝较少"
        return reason, "建议取关"
    return "", "保留"


def autofit(ws, widths=None):
    widths = widths or {}
    for idx, col in enumerate(ws.columns, 1):
        letter = get_column_letter(idx)
        if idx in widths:
            ws.column_dimensions[letter].width = widths[idx]
        else:
            try:
                max_len = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[letter].width = min(max(8, max_len + 2), 40)
            except Exception:
                ws.column_dimensions[letter].width = 12


def style_header(ws, ncols):
    fill = PatternFill("solid", fgColor="4472C4")
    font = Font(color="FFFFFF", bold=True)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "C2"


def safe_sheet_name(name):
    name = re.sub(r"[\\/*?:\[\]]", "", str(name))[:31] or "其他"
    return name


def clean_display_name(name):
    name = str(name or "")
    name = "".join(ch for ch in name if unicodedata.category(ch)[0] != "C" or ch == "\t")
    name = name.strip()
    return name or "（无显示昵称）"


def append_text_row(ws, row):
    ws.append(row)
    for cell in ws[ws.max_row]:
        if isinstance(cell.value, str):
            cell.data_type = "s"


def save_workbook(wb, path):
    try:
        wb.save(path)
        return path
    except PermissionError:
        alt = re.sub(r"\.xlsx$", "", path) + "_" + datetime.datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
        wb.save(alt)
        return alt


def cmd_export(client, args):
    follows = load_json(client.account_file("followings.json"), [])
    analysis = load_json(client.account_file("analysis.json"), {})
    if not follows:
        print("没有关注列表数据，请先运行 fetch。", file=sys.stderr)
        sys.exit(1)
    wb = Workbook()
    wb.remove(wb.active)

    # 使用说明
    ws0 = wb.create_sheet("使用说明")
    lines = [
        "B站关注列表分析报告",
        "",
        "各工作表说明：",
        "全部UP主：所有正常账号的汇总表。",
        "已注销账号：可优先取消关注的账号。",
        "取关建议：根据规则自动给出的建议，你确认后可在“确认取关”列填写“是”，再用 unfollow 命令批量取关。",
        "分类表：按主要分类拆分的表。",
        "",
        "注意：本文件仅保存在你的电脑上，请勿外传。",
    ]
    for i, line in enumerate(lines, 1):
        ws0.cell(row=i, column=1, value=line)

    normal = []
    cancelled = []
    errors = []
    for item in follows:
        mid = str(item.get("mid"))
        rec = analysis.get(mid)
        if not rec:
            continue
        if rec.get("status") == "cancelled":
            cancelled.append(rec)
            continue
        if rec.get("status") == "ok":
            normal.append(rec)
        else:
            errors.append(rec)

    max_top = max((len(r.get("top_videos") or []) for r in normal), default=0)
    main_header = ["UID", "名字", "主要分类", "分类依据", "当前粉丝数", "官方认证", "特别关注",
                   "互相关注", "关注时间", "已关注时长", "最新投稿时间", "距今天数",
                   "最新动态时间", "距今天数(动态)"]
    for i in range(1, max_top + 1):
        main_header.extend([f"热度Top{i}", f"播放量{i}"])

    print("开始生成报告……")

    # 全部UP主
    ws = wb.create_sheet("全部UP主")
    append_text_row(ws, main_header)
    print(f"正在写入“全部UP主”表（{len(normal)} 人）……")
    for idx, rec in enumerate(sorted(normal, key=lambda r: r.get("mtime") or 0, reverse=True), 1):
        top = rec.get("top_videos") or []
        row = [
            rec.get("mid"), clean_display_name(rec.get("uname")), rec.get("category"), rec.get("category_detail"),
            rec.get("fans"), rec.get("official"),
            "是" if rec.get("special") == 1 else "",
            "是" if rec.get("attribute") == 6 else "",
            fmt_ts(rec.get("mtime")),
            human_days(days_since(rec.get("mtime"))),
            fmt_ts(rec.get("latest_video_ts")),
            human_days(days_since(rec.get("latest_video_ts"))),
            fmt_ts(rec.get("latest_dynamic_ts")),
            human_days(days_since(rec.get("latest_dynamic_ts"))),
        ]
        for v in top:
            row.extend([video_label(v), v.get("play")])
        while len(row) < len(main_header):
            row.append("")
        append_text_row(ws, row[:len(main_header)])
        if idx % 500 == 0:
            print(f"  已写入 {idx}/{len(normal)}")
    style_header(ws, len(main_header))
    autofit(ws, {1: 12, 2: 18, 3: 10, 4: 24, 5: 12, 9: 18, 10: 12, 11: 18, 12: 10, 13: 18, 14: 12})

    # 分类表
    by_cat = {}
    for rec in normal:
        by_cat.setdefault(rec.get("category") or "其他/未知", []).append(rec)
    print("正在按分类拆分工作表……")
    for cat in sorted(by_cat, key=lambda c: -len(by_cat[c])):
        recs = by_cat[cat]
        print(f"  分类「{cat}」：{len(recs)} 人")
        ws = wb.create_sheet(safe_sheet_name(cat))
        append_text_row(ws, main_header)
        for rec in sorted(recs, key=lambda r: int(r.get("fans") or 0), reverse=True):
            top = rec.get("top_videos") or []
            row = [
                rec.get("mid"), clean_display_name(rec.get("uname")), rec.get("category"), rec.get("category_detail"),
                rec.get("fans"), rec.get("official"),
                "是" if rec.get("special") == 1 else "",
                "是" if rec.get("attribute") == 6 else "",
                fmt_ts(rec.get("mtime")),
                human_days(days_since(rec.get("mtime"))),
                fmt_ts(rec.get("latest_video_ts")),
                human_days(days_since(rec.get("latest_video_ts"))),
                fmt_ts(rec.get("latest_dynamic_ts")),
                human_days(days_since(rec.get("latest_dynamic_ts"))),
            ]
            for v in top:
                row.extend([video_label(v), v.get("play")])
            while len(row) < len(main_header):
                row.append("")
            append_text_row(ws, row[:len(main_header)])
        style_header(ws, len(main_header))
        autofit(ws, {1: 12, 2: 18, 3: 10, 4: 24, 5: 12, 9: 18, 10: 12, 11: 18, 12: 10, 13: 18, 14: 12})

    # 已注销
    print(f"正在写入已注销账号表（{len(cancelled)} 人）……")
    ws = wb.create_sheet("已注销账号")
    append_text_row(ws, ["UID", "名字", "关注时间", "已关注时长", "备注"])
    for rec in cancelled:
        append_text_row(ws, [rec.get("mid"), clean_display_name(rec.get("uname")) or "账号已注销",
                             fmt_ts(rec.get("mtime")), human_days(days_since(rec.get("mtime"))), "账号已注销"])
    style_header(ws, 5)
    autofit(ws, {1: 12, 2: 20, 3: 18, 4: 12, 5: 20})

    # 分析失败
    if errors:
        print(f"正在写入分析失败表（{len(errors)} 人）……")
        ws = wb.create_sheet("分析失败")
        append_text_row(ws, ["UID", "名字", "关注时间", "已关注时长", "错误信息"])
        for rec in errors:
            append_text_row(ws, [rec.get("mid"), clean_display_name(rec.get("uname")), fmt_ts(rec.get("mtime")),
                                 human_days(days_since(rec.get("mtime"))), rec.get("error")])
        style_header(ws, 5)
        autofit(ws, {1: 12, 2: 20, 3: 18, 4: 12, 5: 40})

    # 取关建议
    print("正在生成取关建议表……")
    plan_header = ["UID", "名字", "当前粉丝数", "互相关注", "特别关注",
                   "关注时间", "已关注时长", "最新动态/视频时间", "距今天数(天)", "未更新时长", "状态",
                   "建议理由", "建议操作", "确认取关(是/否)", "主要分类",
                   "热门视频Top1", "热门视频Top2", "热门视频Top3"]
    plan_rows = []
    for idx, rec in enumerate(sorted(normal + cancelled + errors, key=lambda r: r.get("mtime") or 0, reverse=True), 1):
        if rec.get("status") == "error":
            reason, action = "分析失败，待重试", "保留"
        else:
            reason, action = make_reason(rec, args.days)
        top = rec.get("top_videos") or []
        top_labels = []
        for v in top[:3]:
            top_labels.append(f"{video_label(v)}（播放{v.get('play')}）")
        while len(top_labels) < 3:
            top_labels.append("")
        plan_rows.append([
            rec.get("mid"), clean_display_name(rec.get("uname")),
            rec.get("fans") if rec.get("status") != "cancelled" else "",
            "是" if rec.get("attribute") == 6 else "",
            "是" if rec.get("special") == 1 else "",
            fmt_ts(rec.get("mtime")), human_days(days_since(rec.get("mtime"))),
            fmt_ts(rec.get("latest_post_ts")),
            days_since(rec.get("latest_post_ts")),
            human_days(days_since(rec.get("latest_post_ts"))),
            rec.get("status"), reason, action, "否",
            rec.get("category") if rec.get("status") != "cancelled" else "已注销",
            *top_labels,
        ])
        if idx % 1000 == 0:
            print(f"  已处理 {idx}/{len(normal) + len(cancelled) + len(errors)}")

    ws = wb.create_sheet("取关建议")
    append_text_row(ws, plan_header)
    for row in plan_rows:
        append_text_row(ws, row)
    style_header(ws, len(plan_header))
    autofit(ws, {1: 12, 2: 20, 3: 12, 4: 10, 5: 10, 6: 18, 7: 12, 8: 18, 9: 12, 10: 12, 11: 10, 12: 28, 13: 12, 14: 16, 15: 10, 16: 40, 17: 40, 18: 40})

    action_idx = plan_header.index("建议操作")
    suggested_rows = [row for row in plan_rows if row[action_idx] == "建议取关"]
    ws_suggest = wb.create_sheet("仅建议取关")
    append_text_row(ws_suggest, plan_header)
    for row in suggested_rows:
        append_text_row(ws_suggest, row)
    style_header(ws_suggest, len(plan_header))
    autofit(ws_suggest, {1: 12, 2: 20, 3: 12, 4: 10, 5: 10, 6: 18, 7: 12, 8: 18, 9: 12, 10: 12, 11: 10, 12: 28, 13: 12, 14: 16, 15: 10, 16: 40, 17: 40, 18: 40})
    print(f"建议取关共 {len(suggested_rows)} 人")

    print("正在保存报告文件……")
    report_path = client.account_file("bili_follow_report.xlsx")
    saved_report = save_workbook(wb, report_path)
    print(f"报告已生成：{saved_report}")

    # 生成可编辑的取关计划表（仅含建议取关，便于快速浏览）
    plan_path = client.account_file("unfollow_plan.xlsx")
    if os.path.exists(plan_path) and not getattr(args, "refresh_plan", False):
        print(f"取关计划表已存在，未覆盖：{plan_path}（如需重新生成请加 --refresh-plan）")
    else:
        print("正在生成可编辑取关计划表（仅含建议取关）……")
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "取关计划"
        append_text_row(ws2, plan_header)
        for idx, row in enumerate(suggested_rows, 1):
            append_text_row(ws2, row)
            if idx % 500 == 0:
                print(f"  已写入 {idx}/{len(suggested_rows)}")
        style_header(ws2, len(plan_header))
        autofit(ws2, {1: 12, 2: 20, 3: 12, 4: 10, 5: 10, 6: 18, 7: 12, 8: 18, 9: 12, 10: 12, 11: 10, 12: 28, 13: 12, 14: 16, 15: 10, 16: 40, 17: 40, 18: 40})
        saved_plan = save_workbook(wb2, plan_path)
        print(f"取关计划表已生成：{saved_plan}（仅含建议取关 {len(suggested_rows)} 人）")


def read_plan_targets(client):
    plan_path = client.account_file("unfollow_plan.xlsx")
    if not os.path.exists(plan_path):
        print("未找到取关计划表，请先运行 export。", file=sys.stderr)
        sys.exit(1)
    from openpyxl import load_workbook

    wb = load_workbook(plan_path, data_only=True)
    ws = wb["取关计划"] if "取关计划" in wb.sheetnames else wb.active
    header = [c.value for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}
    uid_col = idx.get("UID", 0)
    confirm_col = idx.get("确认取关(是/否)")
    if confirm_col is None:
        print("计划表缺少“确认取关(是/否)”列。", file=sys.stderr)
        sys.exit(1)
    targets = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[confirm_col] is None:
            continue
        val = str(row[confirm_col]).strip().lower()
        if val in ("是", "yes", "true", "1", "y", "取关", "确定"):
            targets.append(str(row[uid_col]))
    return targets


def read_selected_targets(client):
    targets = load_json(client.account_file("selected_unfollow.json"), [])
    return [str(m) for m in targets]


def run_unfollow(client, targets, delay_min, delay_max):
    state = load_json(client.account_file("unfollow_state.json"), {})
    done = set(state.get("done", []))
    todo = [m for m in targets if m not in done]
    print(f"共需取关 {len(targets)} 人，已完成 {len(done)}，本次待处理 {len(todo)}")
    for i, mid in enumerate(todo):
        print(f"[{i + 1}/{len(todo)}] 取关 {mid}")
        resp = client.unfollow(mid)
        code = resp.get("code")
        if code == 0:
            print("  成功")
            done.add(mid)
        else:
            print(f"  失败 code={code}：{resp.get('message')}")
            if code in (-352, -412):
                print("  遇到风控，请等待一段时间后重新运行本命令（会自动续传）。")
                state["done"] = sorted(done)
                save_json(client.account_file("unfollow_state.json"), state)
                break
        state["done"] = sorted(done)
        save_json(client.account_file("unfollow_state.json"), state)
        time.sleep(delay_min + random.random() * max(0.0, delay_max - delay_min))
    print(f"处理结束，累计完成 {len(done)} 人。")


def cmd_unfollow(client, args):
    ensure_login(client)
    if args.from_selected:
        targets = read_selected_targets(client)
        if not targets:
            print("尚未在“取关管理”中选择取关对象。", file=sys.stderr)
            return
    else:
        targets = read_plan_targets(client)
    if not targets:
        print("计划表中没有标记“是”的账号。")
        return
    run_unfollow(client, targets, args.delay_min, args.delay_max)


def cmd_unfollow_cancelled(client, args):
    ensure_login(client)
    analysis = load_json(client.account_file("analysis.json"), {})
    targets = [mid for mid, rec in analysis.items() if rec.get("status") == "cancelled"]
    if not targets:
        print("没有检测到已注销账号。")
        return
    print(f"检测到 {len(targets)} 个已注销账号，将直接取消关注（请确认这是你要的操作）。")
    if not args.yes:
        try:
            ans = input("确认继续？输入 yes 继续：")
        except (EOFError, OSError):
            print("未收到确认输入，已取消操作。", file=sys.stderr)
            return
        if ans.strip().lower() != "yes":
            print("已取消。")
            return
    run_unfollow(client, targets, args.delay_min, args.delay_max)


def cmd_status(client, args):
    follows = load_json(client.account_file("followings.json"), [])
    analysis = load_json(client.account_file("analysis.json"), {})
    done = sum(1 for r in analysis.values() if r.get("done"))
    cancelled = sum(1 for r in analysis.values() if r.get("status") == "cancelled")
    errors = sum(1 for r in analysis.values() if r.get("status") == "error")
    print(f"关注总数：{len(follows)}")
    print(f"已分析：{done}，已注销：{cancelled}，错误：{errors}")
    print(f"登录 UID：{client.nav_mid()}")


def cmd_incremental_sync(client, args):
    uid = ensure_login(client)
    old = load_json(client.account_file("followings.json"), [])
    old_mids = {str(x.get("mid")) for x in old}
    new_items = client.fetch_followings(uid, delay=args.delay, start_page=1, max_pages=args.max_pages)
    save_json(client.account_file("followings.json"), new_items)
    new_mids = {str(x.get("mid")) for x in new_items}
    removed = old_mids - new_mids
    added = [x for x in new_items if str(x.get("mid")) not in old_mids]

    analysis = load_json(client.account_file("analysis.json"), {})
    for mid in removed:
        analysis.pop(mid, None)

    print(f"关注列表：{len(new_items)} 人；新增 {len(added)}；移除 {len(removed)}")

    # 新增账号完整分析
    for idx, item in enumerate(added, 1):
        rec = analyze_one(client, item, no_dynamic=args.no_dynamic_update, top_n=args.top_n)
        rec["done"] = rec.get("status") != "error"
        rec["updated"] = now_ts()
        analysis[str(item.get("mid"))] = rec
        save_json(client.account_file("analysis.json"), analysis)
        print(f"新增分析 [{idx}/{len(added)}] {rec.get('uname') or rec.get('mid')}：{rec.get('category')}")

    added_mids = {str(x.get("mid")) for x in added}
    total_existing = len(new_items) - len(added)
    updated_count = 0
    if args.no_dynamic_update:
        print("已选择不更新现有账号的最新动态时间。")
    else:
        print(f"开始为现有 {total_existing} 个账号更新最新动态时间……")
        risk_streak = 0
        for idx, item in enumerate(new_items, 1):
            mid = str(item.get("mid"))
            if mid in added_mids:
                continue
            rec = analysis.get(mid)
            if not rec or rec.get("status") != "ok":
                continue
            dyn_ts = None
            dyn_type = None
            dyn_text = ""
            for attempt in range(args.max_risk_retries + 1):
                dyn_ts, dyn_type, dyn_text = client.get_latest_dynamic(mid)
                if dyn_ts is not None or dyn_type not in (-352, -412):
                    break
                print(f"  风控，冷却 {args.cooldown} 秒后重试（第 {attempt + 1} 次）……")
                time.sleep(args.cooldown)
            if dyn_ts:
                rec["latest_dynamic_ts"] = to_int(dyn_ts)
                rec["latest_dynamic_type"] = dyn_type
                rec["latest_dynamic_text"] = dyn_text
            else:
                if dyn_type in (-352, -412):
                    risk_streak += 1
                    if risk_streak >= 3:
                        print("连续遇到风控，已暂停动态更新。请等待更长时间后重新运行 incremental-sync。")
                        break
                else:
                    risk_streak = 0
            candidates = [rec.get("latest_dynamic_ts"), rec.get("latest_video_ts")]
            rec["latest_post_ts"] = max([ts for ts in candidates if ts], default=None)
            rec["updated"] = now_ts()
            analysis[mid] = rec
            updated_count += 1
            if updated_count % 50 == 0:
                save_json(client.account_file("analysis.json"), analysis)
                print(f"  更新动态进度：{updated_count}/{total_existing}")
            lo = args.delay_min if args.delay_min is not None else 0.2
            hi = args.delay_max if args.delay_max is not None else 0.6
            time.sleep(lo + random.random() * max(0.0, hi - lo))

    save_json(client.account_file("analysis.json"), analysis)
    print(f"快速更新完成：新增 {len(added)}，移除 {len(removed)}，更新动态 {updated_count}")
    if args.export:
        args.refresh_plan = True
        cmd_export(client, args)


def cmd_run(client, args):
    uid = ensure_login(client)
    items = client.fetch_followings(uid, delay=args.delay, start_page=1, max_pages=args.max_pages)
    path = client.account_file("followings.json")
    save_json(path, items)
    print(f"已保存关注列表：{len(items)} 人 → {path}")
    cmd_analyze(client, args)
    cmd_export(client, args)


def build_parser():
    p = argparse.ArgumentParser(description="B站关注列表清理与分析工具")
    sub = p.add_subparsers(dest="cmd", required=True)

    sub.add_parser("login", help="扫码登录并保存凭证")
    sub.add_parser("logout", help="退出当前账号登录")

    f = sub.add_parser("fetch", help="拉取完整关注列表")
    f.add_argument("--delay", type=float, default=0.8)
    f.add_argument("--start", type=int, default=1)
    f.add_argument("--max-pages", type=int, default=None)

    a = sub.add_parser("analyze", help="逐个分析关注账号")
    a.add_argument("--start", type=int, default=0)
    a.add_argument("--limit", type=int, default=None)
    a.add_argument("--only-new", action="store_true", help="跳过已完成的分析")
    a.add_argument("--no-dynamic", action="store_true", help="不请求最新动态，只按最新视频判断")
    a.add_argument("--top-n", type=int, default=10, help="分析热度最高的视频数量")
    a.add_argument("--delay-min", type=float, default=1.2)
    a.add_argument("--delay-max", type=float, default=2.5)
    a.add_argument("--cooldown", type=float, default=300, help="遇到风控后的冷却秒数")
    a.add_argument("--max-risk-retries", type=int, default=2, help="单个账号遇到风控时的额外重试次数")

    e = sub.add_parser("export", help="生成 Excel 报告和取关计划表")
    e.add_argument("--days", type=int, default=180, help="多少天未更新即建议取关")
    e.add_argument("--refresh-plan", action="store_true", help="覆盖并重新生成取关计划表")

    u = sub.add_parser("unfollow", help="按计划表批量取关（确认列为“是”）")
    u.add_argument("--delay-min", type=float, default=2.0)
    u.add_argument("--delay-max", type=float, default=4.0)
    u.add_argument("--from-selected", action="store_true", help="取关“取关管理”里保存的选择")

    uc = sub.add_parser("unfollow-cancelled", help="取关所有已注销账号")
    uc.add_argument("--yes", action="store_true", help="跳过二次确认")
    uc.add_argument("--delay-min", type=float, default=2.0)
    uc.add_argument("--delay-max", type=float, default=4.0)

    sub.add_parser("status", help="查看进度")

    q = sub.add_parser("incremental-sync", help="增量同步：移除已取关、新增账号完整分析、现有账号更新最新动态")
    q.add_argument("--delay", type=float, default=0.8)
    q.add_argument("--max-pages", type=int, default=None)
    q.add_argument("--top-n", type=int, default=10)
    q.add_argument("--delay-min", type=float, default=0.8)
    q.add_argument("--delay-max", type=float, default=1.8)
    q.add_argument("--cooldown", type=float, default=300)
    q.add_argument("--max-risk-retries", type=int, default=2)
    q.add_argument("--days", type=int, default=180)
    q.add_argument("--export", action="store_true", help="完成后自动重新生成表格")
    q.add_argument("--no-dynamic-update", action="store_true", help="不更新现有账号的最新动态时间")

    r = sub.add_parser("run", help="依次执行 fetch + analyze + export")
    r.add_argument("--delay", type=float, default=0.8)
    r.add_argument("--start", type=int, default=0)
    r.add_argument("--limit", type=int, default=None)
    r.add_argument("--only-new", action="store_true")
    r.add_argument("--no-dynamic", action="store_true")
    r.add_argument("--top-n", type=int, default=10)
    r.add_argument("--delay-min", type=float, default=1.2)
    r.add_argument("--delay-max", type=float, default=2.5)
    r.add_argument("--cooldown", type=float, default=300)
    r.add_argument("--max-risk-retries", type=int, default=2)
    r.add_argument("--days", type=int, default=180)
    r.add_argument("--max-pages", type=int, default=None)
    return p


def main():
    parser = build_parser()
    args = parser.parse_args()
    client = BiliClient()
    if args.cmd == "login":
        cmd_login(client, args)
    elif args.cmd == "logout":
        cmd_logout(client, args)
    elif args.cmd == "fetch":
        cmd_fetch(client, args)
    elif args.cmd == "analyze":
        cmd_analyze(client, args)
    elif args.cmd == "export":
        cmd_export(client, args)
    elif args.cmd == "unfollow":
        cmd_unfollow(client, args)
    elif args.cmd == "unfollow-cancelled":
        cmd_unfollow_cancelled(client, args)
    elif args.cmd == "status":
        cmd_status(client, args)
    elif args.cmd == "incremental-sync":
        cmd_incremental_sync(client, args)
    elif args.cmd == "run":
        cmd_run(client, args)


if __name__ == "__main__":
    main()
